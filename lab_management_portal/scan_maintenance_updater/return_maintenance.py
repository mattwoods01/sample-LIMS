import requests
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np

FILE_NAME = "Maintenance.xlsx"

EXPORT_COLUMNS = ['Name',
                  'Maintenance Type',
                  'Maintenance Date',
                  'Maintenance Status',
                  'Previous Maintenance',
                  'Owner',
                  'Primary Contact',
                  'Lab Manager Owner',
                  'Secondary Contact',
                  'Cost Center',
                  'Location',
                  'Equipment Type',
                  'Equipment Description',
                  'Model Number',
                  'Serial Number',
                  'ULS Tag',
                  'Inventory Tag Number',
                  'Barcode',
                  'Purchase Date',
                  'Warranty Expiration Date',
                  'URL',
                  'Equipment ID'
                  ]

MAIN_TYPES = ["Calibration",
              "Preventative Maintenance",
              "IQ",
              "OQ",
              "PQ",
              "Defrost"
              ]

def format_email_list(compiled_df):
    
    """
    format email list based off labguru primary and secondary contact on equipments
    """
    manager_df = compiled_df.drop_duplicates(subset='Primary Contact')
    manager_df = manager_df.dropna(subset=['Primary Contact'])
    manager_ls = manager_df['Primary Contact'].tolist()

    owner_df = compiled_df.drop_duplicates(subset='Secondary Contact')
    owner_df = owner_df.dropna(subset=['Secondary Contact'])
    owner_ls = owner_df['Secondary Contact'].tolist()

    input_list = owner_ls + manager_ls
    unique_list = list(set(input_list))

    while "" in unique_list:
        unique_list.remove("")

    email_list = ', '.join(unique_list)
    
    tr = sys.argv[1]  
    store_variable("calibration_list", {'tmp_path': f'{tr}/rendered_files/{FILE_NAME}'})
    store_variable("email_list", str(email_list))

class MaintenanceOutputCreator:
  def __init__(self, folder_name):
    headers = {'Content-type': 'application/json'}
    payload = {'token': "5223e9371b618b588382bfa4daab92af1f8fcf32"}
    self.base_url = 'https://thermofisher.labguru.com'
    self.today = datetime.today().date()
    self.current_year = self.today.year
    self.current_month = self.today.month
    self.next_month_date = datetime(self.current_year, self.current_month, 1) + relativedelta(months=1)
    self.next_month = self.next_month_date.month
    self.instrument_catalogue = requests.get(self.base_url + '/api/v1/instruments', json=payload, headers=headers).text
    self.compiled_ls = [] 
    self.directory = f"{folder_name}/{FILE_NAME}"

  def generate_maintenance_sheet(self):
    """
    generate maintenance sheet from information from labguru
    """

    instrument_catalogue_df = pd.read_json(self.instrument_catalogue, dtype=False)
    
    for i, item in instrument_catalogue_df.iterrows():      
      main_types_ls = []

      for main_type in MAIN_TYPES:
        main_types_ls.append(str(item[f"{main_type} Frequency (months)"]) + '_' + main_type + '_' + str(item[f"{main_type} Date (YYYY-MM-DD)"]))

      main_types_df = pd.DataFrame(main_types_ls, columns=['data'])
      main_types_ls = list(main_types_df['data'].apply(lambda x: tuple(x.split("_"))))
      prev_maintenance_df = pd.DataFrame(main_types_ls, columns=['months', "name", "date"])

      prev_maintenance_df.replace(['', 'nan'], np.nan, inplace=True)
      prev_maintenance_df = prev_maintenance_df.dropna()

      for j, row in prev_maintenance_df.iterrows():
            maintenance_date_str = str(row['date'])
            maintenance_name = str(row['name'])

            try:
              maintenance_frequency = int(row['months'])

            except:
              maintenance_frequency = 0
              maintenance_date_str = "1900-01-01"
              
            if self.current_month == 12:
              next_month = 1
              next_year = self.current_year + 1

            else:
              next_month = self.current_month + 1
              next_year = self.current_year

            if maintenance_date_str == "":
              maintenance_date = datetime.strptime("01/01/1900", "%m/%d/%Y")

            else:
              maintenance_date = datetime.strptime(maintenance_date_str, '%Y-%m-%d')

              updated_maintenance_date = maintenance_date + relativedelta(months=maintenance_frequency)
              maintenance_year = updated_maintenance_date.year
              maintenance_month = updated_maintenance_date.month
              updated_maintenance_date_str = updated_maintenance_date.strftime('%Y-%m-%d')

              indv_pipette_export = [
                    item['name'],
                    maintenance_name,
                    updated_maintenance_date_str,
                    maintenance_date_str,
                    item['owner_name'],
                    item['Primary Contact'],
                    item['Lab Manager Owner'],
                    item['Secondary Contact'],       
                    item['Cost Center'],
                    item['location'],
                    item['equipment_type'],
                    item['Equipment Description'],
                    item['model_number'],
                    item['serial_number'],
                    item['ULS Tag'],
                    item['Inventory Tag Number'],
                    item['barcode'],
                    item['purchase_date'],
                    item['warranty_expired'],
                    self.base_url + item['url'],
                    item['id'],
                ]

              if self.current_year == maintenance_year and self.current_month == maintenance_month:
                indv_pipette_export.insert(3, "Current Month")
                self.compiled_ls.append(indv_pipette_export)

              elif maintenance_year == next_year and maintenance_month == next_month:
                indv_pipette_export.insert(3, "Next Month")
                self.compiled_ls.append(indv_pipette_export)

              elif self.today >= updated_maintenance_date.date():
                indv_pipette_export.insert(3, "Overdue")
                self.compiled_ls.append(indv_pipette_export)
             
  def format_maintenance_excel(self):
    """
    format maintenance excel after it is created
    """
    self.compiled_df = pd.DataFrame(self.compiled_ls, columns=EXPORT_COLUMNS)
    self.compiled_df = self.compiled_df.sort_values(by='Maintenance Status', ascending=True)
    
    with pd.ExcelWriter(self.directory) as writer:
      self.compiled_df.to_excel(writer, sheet_name='Maintenances', index=False)

    workbook = load_workbook(self.directory)
    worksheet = workbook['Maintenances']

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    orange_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='12b150', end_color='12b150', fill_type='solid')

    for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
      for cell in row: 
        if cell.value in ['Name', 'Maintenance Type', 'Maintenance Date']:
          cell.fill = green_fill

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=4):
      cell = row[0]
      if "Overdue" in cell.value:
        for cell_in_row in row:
          cell_in_row.fill = red_fill
  
      elif "Next Month" in cell.value:
        for cell_in_row in row:
          cell_in_row.fill = orange_fill
  
      elif "Current Month" in cell.value:
        for cell_in_row in row:
          cell_in_row.fill = yellow_fill
    
    workbook.save(self.directory)

# maintenance_pipeline = MaintenanceOutputCreator('test')
# maintenance_pipeline.generate_maintenance_sheet()
# maintenance_pipeline.format_maintenance_excel()
