from channels.generic.websocket import AsyncWebsocketConsumer
import os
import random
import pandas as pd
from .sto import sto_generator
from .cherry import InputIter, plate_map
from .bio import BioGenerator
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl import load_workbook
from PyPDF4 import PdfFileWriter, PdfFileReader
import numpy as np
import asyncio
import base64
import json
import helper
import zipfile
from asgiref.sync import sync_to_async

def format_excel(input_workbook, bio_df):
   """
   Formats already created excel to match bioinformatics sheet
   """
   
   wb = load_workbook(input_workbook)
   sheet = wb['Sheet1']
   bioinformatic_header_color = ['A1', 'B1', 'C1', 'D1',
                                 'E1', 'F1', 'G1', 'H1',
                                 'I1', 'J1', 'k1', 'L1',
                                 'M1', 'N1', 'O1', 'P1',
                                 'Q1', 'R1', 'S1', 'T1']
   
   for g, item in enumerate(bioinformatic_header_color):
       sheet[item].fill = PatternFill(start_color='CC3300',
                                      end_color='CC3300',
                                      fill_type='solid')
       sheet[item].font = Font(color='FFFFFFFF')
       sheet[item].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
       side = Side(border_style=None)
       sheet[item].border = Border(left=side, right=side, top=side, bottom=side)

   for k in range(2, len(bio_df) + 2):
       for f in range(1, 21):
           cellref = sheet.cell(row=k, column=f)
           cellref.alignment = Alignment(horizontal='left', vertical='center')
   table = Table(displayName="Table1",
                 ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
   
   sheet.add_table(table)
   wb.save(input_workbook)
   wb.close()

def put_watermark(input_pdf, output_pdf, watermark):
   """
   Places watermark on COAs
   """
   watermark_instance = PdfFileReader(watermark)
   watermark_page = watermark_instance.getPage(0)
   pdf_reader = PdfFileReader(input_pdf)
   pdf_writer = PdfFileWriter()

   for i in range(pdf_reader.getNumPages()):
       page = pdf_reader.getPage(i)
       page.mergePage(watermark_page)
       pdf_writer.addPage(page)

   with open(output_pdf, 'wb') as out:
       pdf_writer.write(out)

class MyConsumer(AsyncWebsocketConsumer):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.folder_name = str(random.randint(1, 1000000))

    async def connect(self):
        """
        Connects to webpage and initializes database
        """
        
        await self.accept()
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Client Connected"}))
        await self.send(text_data=json.dumps({"folder_name": self.folder_name, "event": "folder_name"}))
        os.mkdir(self.folder_name)

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Client Connected"}))
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Querying Lenti Catalogue..."}))

        sql_init = helper.sql_connector()
        sql_data, sql_headers = await sync_to_async(sql_init.sql_connect_and_query)("SELECT * FROM lenti_catalogue;")

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Fetching rows..."}))

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Completed fetch..."}))

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Caching Catalogue..."}))

        self.data_df = pd.DataFrame(sql_data, columns=sql_headers)

        print(self.data_df)

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Caching complete"}))
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Lentiarray Catalogue ready for viewing"}))
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Cherrypick Tool ready to use"}))

        print("Connected")

    async def disconnect(self, close_code):

        """
        Disconnects from server and removes folder
        """
        helper.remove_folder(self.folder_name)

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Client Disconnected"}))
        print("Disconnected")

    async def receive(self, text_data=None):

        """
        receives events from webpage inputs
        """
        data = json.loads(text_data)
        event = data.get("event")
        print(event)

        if event == "run":
            await self.run_program(data)
        elif event == 'update_csv':
            await self.update_csv(data)
        elif event == 'file_upload':
            await self.file_upload(data)
        
    async def run_program(self, data):
        """
        runs entire pipeline
        """
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Cherrypick initiated..."}))
        data = data['data']
        print(data)
    
        form_headers = data['form_headers']
        form_data = data['form_data']
        pref_headers = data['pref_headers']
        pref_data = data['pref_data']
        sto_head = data['sto_headers']
        sto_data = data['sto_data']
        form_rep = int(data['form_rep'])
        check_bio = bool(data['check_bio'])
        check_coa = bool(data['check_sto'])

        # folder_name = cache.get(folder_cache)

        form_df = pd.DataFrame(form_data, columns=form_headers)
        # data_df = cache.get(sql_cache)

        export_df = pd.DataFrame()
        non_existent_df = pd.DataFrame()
        bypassed_df = pd.DataFrame()
        multiple_df = pd.DataFrame()
        pref_library = pd.DataFrame(pref_data, columns=pref_headers)

        cherry_pick = InputIter(self.data_df, form_rep, pref_library)

        form_df.replace(['', 'nan'], np.nan, inplace=True)
        form_df = form_df.dropna()

        await asyncio.sleep(0.5)
        await self.send(text_data=json.dumps({"event": "return_output", "message": f"Begin Cherrypick matching..."}))
        for i, entry in form_df.iterrows():
            entry_df = pd.DataFrame(entry).T.reset_index()
            well_position, plate_position = str(entry['Well Position']), int(entry['Plate'])

            if entry['ARRAY OR POOL'] == 'ARRAY': 
                output, all_matches, bypassed, non_existant = cherry_pick.find_exact_match(entry_df,
                                                                                                        'crispr_id',
                                                                                                        ('B', 'P')) 
                form_type = 'ARRAY'
            
            elif entry['ARRAY OR POOL'] == 'POOL':
                output, all_matches, bypassed, non_existant = cherry_pick.find_exact_match(entry_df,
                                                                                                        'ncbi_gene',
                                                                                                        ('B', 'M'))
                form_type = 'POOL'
            
            await asyncio.sleep(0.25)
            multiple_df = pd.concat([multiple_df, all_matches], axis=0)
            
            if non_existant.empty is False:
                await self.send(text_data=json.dumps({"event": "return_output", "message": f"Non-existant found for {entry['CRISPR ID OR ENTREZ ID']} - SKIPPING"}))
                non_existent_df = pd.concat([non_existent_df, non_existant], axis=0)

            else:
                if bypassed.empty is False:
                    await self.send(text_data=json.dumps({"event": "return_output", "message": f"Bypass found for {entry['CRISPR ID OR ENTREZ ID']}"}))
                    bypassed_df = pd.concat([bypassed_df, bypassed], axis=0)

                if len(bypassed.index) == 1 and form_type == 'ARRAY':
                    await self.send(text_data=json.dumps({"event": "return_output", "message": f"Single Bypass found for {entry['CRISPR ID OR ENTREZ ID']} - SKIPPING"}))

                else:
                    output = pd.concat([entry_df, output], axis=1)
                    output.set_index('index', inplace=True)
                    await self.send(text_data=json.dumps({"event": "return_output", "message": f"Match found for {entry['CRISPR ID OR ENTREZ ID']}"}))
                    output['destination barcode'], output['destination well'] = plate_position, well_position
                    export_df = pd.concat([export_df, output], axis=0, ignore_index=True) 
                    print(export_df)
         
        await self.send(text_data=json.dumps({"event": "return_output", "message": f"Finished cherrypick matching"})) 
        await asyncio.sleep(0.5)
        await self.send(text_data=json.dumps({"event": "return_output", "message": f"Finding plate using R{form_rep}..."}))
        cherrypick_list = cherry_pick.find_plate_by_replicate(export_df)
        plate_map_output = plate_map(export_df)
        await self.send(text_data=json.dumps({"event": "return_output", "message": f"Creating plate map..."}))
        await asyncio.sleep(0.5)

        await self.send(text_data=json.dumps({"event": "return_output", "message": f"Creating excel files..."}))
        multiple_df.to_excel(f"{self.folder_name}/all possible matches.xlsx", index=False)
        non_existent_df.to_excel(f"{self.folder_name}/Missing targets.xlsx", index=False)
        plate_map_output.to_excel(f'{self.folder_name}/Plate map.xlsx', index=False)
        cherrypick_list.to_csv(f"{self.folder_name}/Cherrypick.csv", index=False)
        export_df.to_excel(f"{self.folder_name}/Cherrypick.xlsx", index=False)
        bypassed_df.to_excel(f"{self.folder_name}/Bypassed.xlsx", index=False)
        await self.send(text_data=json.dumps({"event": "return_output", "message": f"Finished creating excel files..."}))
        await asyncio.sleep(0.5)

        full_cherrypick = pd.concat(([cherrypick_list, export_df]), axis=1)
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Finsihed generating cherrypick file"}))

        if check_bio == True:
            await self.send(text_data=json.dumps({"event": "return_output", "message": "Generating bioinformatics file..."}))
            await asyncio.sleep(0.5)

            bio_gen = BioGenerator(self.data_df)
            for i, entry in full_cherrypick.iterrows():
                try:
                    bio_gen.create_bio(entry)
                
                except Exception as e:
                    await self.send(text_data=json.dumps({"event": "return_output", "message": f"{str(e)}"}))

                await self.send(text_data=json.dumps({"event": "return_output", "message": f"Generated Export for {entry['CRISPR ID OR ENTREZ ID']}"}))
                await asyncio.sleep(0.5)

            await self.send(text_data=json.dumps({"event": "return_output", "message": "Finished generating bioinformatics file"}))
            await asyncio.sleep(0.5)

            bio_df, preformatted_df = bio_gen.format_bio()
            bio_df.to_excel(f"{self.folder_name}/Bioinformatics.xlsx", index=False)
            format_excel(f"{self.folder_name}/Bioinformatics.xlsx", bio_df)
            await self.send(text_data=json.dumps({"event": "return_output", "message": "Formatting bioinformatics file..."}))
            await asyncio.sleep(0.5)
            await self.send(text_data=json.dumps({"event": "return_output", "message": "Finsihed formatting bioinformatics file"}))

        if check_coa == True:
            await self.send(text_data=json.dumps({"event": "return_output", "message": "Generating COA file..."}))
            await asyncio.sleep(0.5)

            print(preformatted_df)

            project_groups = preformatted_df.groupby('Library Name')

            for project_name, project_group in project_groups:
                await self.send(text_data=json.dumps({"event": "return_output", "message": f"Creating COA(s) for project: {project_name}"}))
                well_groups = project_group.groupby('destination well')
                well_headers = project_group.columns.tolist()
                sto_init = sto_generator(self.folder_name, well_headers, sto_head, sto_data[int(project_name)-1])

                for well_name, well_group in well_groups:
                    print(str(well_group['Library Type']))

                    if str(well_group['Library Type'].iloc[0]) == 'ARRAY':
                        await self.send(text_data=json.dumps({"event": "return_output", "message": f"Creating single STO COA(s) for {well_name}"}))
                        well_ls = well_group.values.flatten().tolist()
                        file_name = sto_init.generate_single(well_ls)
                        sto_pdf = "lenti_cherrypick/overlays/sto_single.pdf"

                    elif str(well_group['Library Type'].iloc[0]) == 'POOL':
                        await self.send(text_data=json.dumps({"event": "return_output", "message": f"Creating pooled STO COA(s) for {well_name}"}))
                        well_group.reset_index(inplace=True)
                        file_name = sto_init.generate_pool(well_group)
                        sto_pdf = "lenti_cherrypick/overlays/sto_pool.pdf"

                    await asyncio.sleep(0.5)
                    await self.send(text_data=json.dumps({"event": "return_output", "message": "Watermarking PDF..."}))
                    put_watermark(file_name, file_name, sto_pdf)
                    await self.send(text_data=json.dumps({"event": "return_output", "message": "Finsihed watermarking PDF..."}))

            await self.send(text_data=json.dumps({"event": "return_output", "message": "Finished generating COA file..."}))

        await asyncio.sleep(0.5)
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Zipping cherrypick files..."}))
        with zipfile.ZipFile(f'{self.folder_name}.zip', 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(f'{self.folder_name}/'):
                for file in files:
                    zipf.write(f'{self.folder_name}/' + file)

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Finished zipping cherrypick files..."}))
        await asyncio.sleep(0.5)

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Encoding cherrypick files..."}))
        with open(f'{self.folder_name}.zip', 'rb') as file:
            zip_data = file.read()
            encoded_zip_data = base64.b64encode(zip_data).decode('utf-8')
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Finished encoding cherrypick files..."}))
        await asyncio.sleep(0.5)

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Sending cherrypick files..."}))
        await asyncio.sleep(0.5)
        await self.send(text_data=json.dumps({"event": "return_zip", "zip": encoded_zip_data}))
        await asyncio.sleep(0.5)

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Finished sending cherrypick files..."}))
        os.remove(f'{self.folder_name}.zip')
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Cleaning up..."}))
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Finished"}))

    async def update_csv(self, data):
        """
        updates csv for library preferences from webpage
        """
        pref_head = data['pref_head']
        pref_data = data['pref_data']

        pref_df = pd.DataFrame(pref_data, columns=pref_head)
        pref_df.to_csv('lenti_cherrypick/library_pref.csv', index=False, sep=',', header=True, encoding='utf-8')
        await self.send(text_data=json.dumps({"event": "return_pref", "message": {"up_data": pref_data, "up_headers": pref_head}})) 
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Finished updating library preferences file"}))

    async def file_upload(self, data):

        """
        uploads file to client
        """
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Uploading Cherrypick input file"}))
        file_data = data['file_data']
        file_name = data['filename']

        initial_headers = pd.read_csv('lenti_cherrypick/cherry_headers.csv', header=0).columns.tolist()

        await self.send(text_data=json.dumps({"event": "return_output", "message": "Creating Folder for uploaded file"}))
        directory_name = f'{self.folder_name}/{file_name}'

        try:
            # Convert list to bytes
            file_data_bytes = bytes(file_data)

            with open(directory_name, "wb") as file:
                file.write(file_data_bytes)
    
            await self.send(text_data=json.dumps({"event": "return_output", "message": f"File saved to {directory_name}"}))
    
        except Exception as e:
            await self.send(text_data=json.dumps({"event": "return_output", "message": f"Error saving file: {str(e)}"}))
    
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Reading excel..."}))
        input_dataframe = pd.read_excel(directory_name)
        input_dataframe = input_dataframe.astype(str)

        uploaded_data = input_dataframe.values.tolist()
        # uploaded_headers = input_dataframe.columns.tolist()

        await self.send(text_data=json.dumps({"up_data": uploaded_data, "up_headers": initial_headers, "event": "return_data"}))     
        await self.send(text_data=json.dumps({"event": "folder_name", "message": {"folder_name": self.folder_name}}))
        await self.send(text_data=json.dumps({"event": "return_output", "message": "Finished uploading file"})) 